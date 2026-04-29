from __future__ import annotations

import calendar
import math
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd

TIME_RE = re.compile(r"(\d{1,2})(?::(\d{2}))?\s*([ap])\s*-\s*(\d{1,2})(?::(\d{2}))?\s*([ap])", re.I)
MONTH_RE = re.compile(r"^[A-Za-z]{3}-\d{2}$")
ANCHOR = date(2023, 1, 1)
COLUMNS = [
    "person", "date", "dow", "pp", "pp_start", "pp_end", "month", "raw", "type",
    "parsed_hours", "included_hours", "shift_count", "backup_count", "weekend", "night", "source_row",
]


def _minutes(h: str, minute: str | None, ap: str) -> int:
    hour = int(h)
    minute_i = int(minute or 0)
    ap = ap.lower()
    if ap == "a" and hour == 12:
        hour = 0
    if ap == "p" and hour != 12:
        hour += 12
    return hour * 60 + minute_i


def parse_hours(text: str) -> float | None:
    s = str(text).replace("–", "-").replace("—", "-")
    if "backup" in s.lower():
        return 8.0
    m = TIME_RE.search(s)
    if not m:
        return None
    start = _minutes(m.group(1), m.group(2), m.group(3))
    end = _minutes(m.group(4), m.group(5), m.group(6))
    if end <= start:
        end += 24 * 60
    return round((end - start) / 60, 2)


def shift_type(text: str) -> str:
    l = str(text).strip().lower()
    if "backup" in l:
        return "Backup"
    if "night" in l:
        return "Night"
    if "swing" in l:
        return "Swing"
    if "ed mid" in l or "ft mid" in l:
        return "Mid"
    if "tec" in l:
        return "TEC"
    if "orientation" in l:
        return "Orientation"
    if l.startswith("off") or "admin" in l:
        return "Admin-Off"
    if "day" in l or "am" in l or "wknd" in l:
        return "Day"
    return "Other"


def pay_period_for(d: date) -> tuple[str, date, date]:
    idx = (d - ANCHOR).days // 14
    start = ANCHOR + timedelta(days=idx * 14)
    end = start + timedelta(days=13)
    y = start.year
    first = date(y, 1, 1)
    first_idx = math.ceil(((first - ANCHOR).days) / 14)
    first_start = ANCHOR + timedelta(days=first_idx * 14)
    if start < first_start:
        y -= 1
        first = date(y, 1, 1)
        first_idx = math.ceil(((first - ANCHOR).days) / 14)
        first_start = ANCHOR + timedelta(days=first_idx * 14)
    pp_num = (start - first_start).days // 14 + 1
    return f"{y} PP{pp_num:02d}", start, end


def build_date_map(ws) -> dict[int, date]:
    months: list[tuple[int, str]] = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(5, col).value
        if v and MONTH_RE.match(str(v).strip()):
            months.append((col, str(v).strip()))
    col_date: dict[int, date] = {}
    for idx, (start_col, label) in enumerate(months):
        end_col = months[idx + 1][0] - 1 if idx + 1 < len(months) else ws.max_column
        dt = datetime.strptime(label, "%b-%y")
        last = calendar.monthrange(dt.year, dt.month)[1]
        for col in range(start_col, min(end_col, start_col + last - 1) + 1):
            day = ws.cell(6, col).value
            if isinstance(day, (int, float)) and 1 <= int(day) <= last:
                col_date[col] = date(dt.year, dt.month, int(day))
    return col_date


def _looks_like_shift(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return bool(TIME_RE.search(s) or "backup" in s.lower() or "admin" in s.lower())


def find_staff_rows(ws) -> dict[str, int]:
    date_map = build_date_map(ws)
    rows: dict[str, int] = {}
    for row in range(8, min(ws.max_row, 360) + 1):
        name = ws.cell(row, 1).value
        if not name:
            continue
        name_s = str(name).strip()
        if not re.match(r"^[A-Za-z][A-Za-z .'-]{1,40}$", name_s):
            continue
        if name_s.lower() in {"day", "night", "swing", "backup"}:
            continue
        # Include named rows even if first visible months are blank; require either row or continuation has shifts somewhere.
        count = 0
        for rr in (row, row + 1):
            for col in date_map:
                if _looks_like_shift(ws.cell(rr, col).value):
                    count += 1
                    break
        if count:
            rows[name_s] = row
    return rows


def extract_records(workbook_file, selected_staff: Iterable[str] | None = None) -> pd.DataFrame:
    wb = openpyxl.load_workbook(workbook_file, data_only=True, read_only=False)
    ws = wb.active
    date_map = build_date_map(ws)
    staff_rows = find_staff_rows(ws)
    selected = set(selected_staff or staff_rows.keys())
    records = []
    for person, row in staff_rows.items():
        if person not in selected:
            continue
        for rr in (row, row + 1):
            for col, d in sorted(date_map.items(), key=lambda item: item[1]):
                val = ws.cell(rr, col).value
                if val is None or str(val).strip() == "":
                    continue
                raw = str(val).strip()
                if rr == row + 1 and not _looks_like_shift(raw):
                    continue
                typ = shift_type(raw)
                parsed = parse_hours(raw)
                if parsed is None:
                    parsed = 0.0
                excluded = typ == "Admin-Off"
                included_hours = 0.0 if excluded or typ == "Backup" else parsed
                shift_count = 0 if excluded else 1
                label, pp_start, pp_end = pay_period_for(d)
                records.append({
                    "person": person,
                    "date": d,
                    "dow": d.strftime("%a"),
                    "pp": label,
                    "pp_start": pp_start,
                    "pp_end": pp_end,
                    "month": d.strftime("%Y-%m"),
                    "raw": raw,
                    "type": typ,
                    "parsed_hours": float(parsed),
                    "included_hours": float(included_hours),
                    "shift_count": int(shift_count),
                    "backup_count": int(typ == "Backup" and shift_count == 1),
                    "weekend": int(d.weekday() >= 5 and shift_count == 1),
                    "night": int(typ == "Night" and shift_count == 1),
                    "source_row": rr,
                })
    return pd.DataFrame(records, columns=COLUMNS)


def aggregate_pay_periods(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["person", "pp", "pp_start", "pp_end", "shifts", "hours", "backups", "weekends", "nights"])
    out = df.groupby(["person", "pp", "pp_start", "pp_end"], as_index=False).agg(
        shifts=("shift_count", "sum"), hours=("included_hours", "sum"), backups=("backup_count", "sum"), weekends=("weekend", "sum"), nights=("night", "sum")
    )
    out = out[out["shifts"] > 0].sort_values(["pp_start", "person"])
    out["hours"] = out["hours"].round(2)
    return out


def aggregate_monthly(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["person", "month", "shifts", "hours", "backups", "weekends", "nights"])
    out = df.groupby(["person", "month"], as_index=False).agg(
        shifts=("shift_count", "sum"), hours=("included_hours", "sum"), backups=("backup_count", "sum"), weekends=("weekend", "sum"), nights=("night", "sum")
    )
    out = out[out["shifts"] > 0].sort_values(["month", "person"])
    out["hours"] = out["hours"].round(2)
    return out


def comparison_summary(pp_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for person, g in pp_df.groupby("person"):
        rows.append({
            "person": person,
            "pay_periods": int(len(g)),
            "total_shifts": int(g["shifts"].sum()),
            "total_hours": round(float(g["hours"].sum()), 1),
            "avg_hours_pp": round(float(g["hours"].mean()), 1),
            "median_hours_pp": round(float(g["hours"].median()), 1),
            "max_hours_pp": round(float(g["hours"].max()), 1),
            "pp_ge_80": int((g["hours"] >= 80).sum()),
            "pp_ge_64": int((g["hours"] >= 64).sum()),
            "pp_ge_56": int((g["hours"] >= 56).sum()),
            "total_backups": int(g["backups"].sum()),
            "weekend_shifts": int(g["weekends"].sum()),
            "night_shifts": int(g["nights"].sum()),
        })
    return pd.DataFrame(rows).sort_values("person") if rows else pd.DataFrame()


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")
